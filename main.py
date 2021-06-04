import os
import shutil
import openpyxl
import pandas as pd

from datetime import datetime
from os import path
from openpyxl import load_workbook
from Utils import Utils
from openpyxl.styles.borders import Border, Side

from openpyxl.styles import PatternFill
from openpyxl.styles.colors import WHITE
from openpyxl.styles import Border, Side



def duplicateTemplateLTGC(tempLTGC_Path, out, compCode):
    companyDir = out + "/" + compCode

    # creating new DIR base on company code
    if not path.exists(out + "/" + compCode):
        os.mkdir(os.path.join(out, compCode))

    shutil.copy(tempLTGC_Path,
                companyDir + "/LTGC_CEIRMasterlist_ExtraCols_" + compCode + ".xlsx")

    return companyDir + "/LTGC_CEIRMasterlist_ExtraCols_" + compCode + ".xlsx"


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    pass


def getData(inFile_LTGC, outPath):
    print("============================")
    print("Starting LTGC Files")
    print("============================")
    util = Utils()
    df = pd.read_excel(inFile_LTGC, sheet_name='Eligible Population',
                       header=1, dtype={'PhilHealth_ID*': str, 'Contact_number_of_employer*': str,
                                        'Contact_No.*': str, 'Age': str}, na_filter=False)

    groups = df.groupby('Company')

    for i, comp in groups:
        # companyCode = companyNameLookUp(i)
        companyCode = util.companyNameLookUpMethod(i)

        comp = comp.astype(str)

        # get num rows
        numrows = len(comp.index)

        print(i + ' (' + companyCode + ") has " + str(numrows) + " records", end='')

        templateFile = util.duplicateTemplateLTGC(tempLTGC_Path, outPath, companyCode, i)

        # border settings
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        theFile = openpyxl.load_workbook(templateFile)
        currentSheet = theFile["Eligible Population"]
        util.addingDataValidation(currentSheet, numrows)

        # set cell border: has 75 cols
        # for row in range(2, numrows + 3):
        #     for col in range(1, 77):
        #         currentSheet.cell(row=row, column=col).border = thin_border

        set_border(currentSheet, "A3:BX" + str(numrows + 2))

        ## Set bg color in a cell
        # currentSheet.cell(row=3, column=33).fill = PatternFill(start_color="ffffff", fill_type = "solid")

        theFile.save(templateFile)

        writer = pd.ExcelWriter(templateFile, engine='openpyxl', mode='a')
        writer.book = load_workbook(templateFile)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        comp.to_excel(writer, "Eligible Population", startrow=2, header=False, index=False)
        writer.save()

        print(".....Done!")

        # print(comp)

    pass


if __name__ == '__main__':
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    pd.set_option('display.width', None)

    today = datetime.today()
    dateTime = today.strftime("%m_%d_%y_%H%M%S")

    inPath = r"/Users/ranperedo/Documents/Vaccine/LTGSplit/in"
    outPath = r"/Users/ranperedo/Documents/Vaccine/LTGSplit/out/ltgc"
    templateFilePath = r"/Users/ranperedo/Documents/Vaccine/LTGSplit/template"

    inFile_LTGC = inPath + "/LTGC_CEIRMasterlist_Combined.xlsx"
    # inFile_LTGC = inPath + "/LTGC_CEIRMasterlist.xlsx"

    tempLTGC_Path = templateFilePath + "/LTGC_CEIRMasterlist_ExtraCols.xlsx"

    # Excel Templates: create copy

    print("Split File Script......")

    if path.exists(inFile_LTGC) and path.isfile(inFile_LTGC):
        getData(inFile_LTGC, outPath)
    else:
        print(str(inFile_LTGC) + " File is invalid or does not exist")
